# Code to create an Excel file with informations about a particular stock
from turtle import st
from matplotlib.axis import Tick
import yfinance
import openpyxl
import pandas as pd
import numpy as np
from openpyxl import Workbook
from datetime import datetime, timedelta    
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

Ticker = 'META'
benchmark = "^GSPC"
rf = 0.02 # risk free rate (Valor em %)
stock = yfinance.Ticker(Ticker)



def calcular_beta(stock, benchmark):
    stock = yfinance.Ticker(stock)
    benchmark = yfinance.Ticker(benchmark)
    # 5 years of data
    stock_data = stock.history(period="5y")
    benchmark_data = benchmark.history(period="5y")
    # Calculate returns
    stock_returns = stock_data["Close"].pct_change().dropna()
    benchmark_returns = benchmark_data["Close"].pct_change().dropna()
    # Calculate covariance of stock vs benchmark
    covariance = stock_returns.cov(benchmark_returns)
    # Calculate variance of benchmark
    variance = benchmark_returns.var()
    # Calculate beta
    beta = covariance / variance
    return beta


def calcular_VaR(stock, period='5y', interval='1d', confidence_level=0.95):
    """
    Calcula o Value at Risk (VaR) diário a 95% de um ativo.

    :param stock: String, o ticker do ativo no formato usado pelo Yahoo Finance.
    :param period: String, o período de tempo para coleta de dados históricos (ex: '1y', '5y').
    :param interval: String, o intervalo dos dados (ex: '1d' para diário, '1wk' para semanal).
    :param confidence_level: Float, o nível de confiança para o cálculo do VaR (ex: 0.95 para 95%).
    :return: Float, o VaR diário a 95%.
    """
    try:
        # Carrega os dados históricos do ativo
        stock_data = yfinance.download(stock, period=period, interval=interval)['Close'].pct_change().dropna()

        # Calcula o VaR
        VaR = -np.percentile(stock_data, (1 - confidence_level) * 100)
        return VaR
    except Exception as e:
        print(f"Erro: {e}")
        return None



def calcular_sharpe_ratio(ticker, rf):
    """
    Calcula o Sharpe Ratio para um dado ativo de um ano atrás até hoje.

    :param ticker: String, o ticker do ativo no formato usado pelo Yahoo Finance.
    :param rf: Float, a taxa de retorno livre de risco anual.
    :return: Float, o Sharpe Ratio.
    """
    # Determina a data atual e a data de um ano atrás
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365)

    # Formata as datas para o formato requerido pelo yfinance
    start_date = start_date.strftime('%Y-%m-%d')
    end_date = end_date.strftime('%Y-%m-%d')

    # Carrega os dados históricos do ativo
    data = yfinance.download(ticker, start=start_date, end=end_date)

    # Calcula os retornos diários
    data['Daily Return'] = data['Adj Close'].pct_change()

    # Calcula o retorno médio diário e o desvio padrão dos retornos diários
    mean_daily_return = data['Daily Return'].mean()
    std_daily_return = data['Daily Return'].std()

    # Converte a taxa livre de risco em base diária
    rf_daily = (1 + rf) ** (1/252) - 1

    # Calcula o Sharpe Ratio anualizado
    sharpe_ratio = (mean_daily_return - rf_daily) / std_daily_return
    sharpe_ratio_anualizado = sharpe_ratio * np.sqrt(252)

    return sharpe_ratio_anualizado

def obter_preco_atual(ticker):
    """
    Obtem o preço atual de uma ação.

    :param ticker: String, o ticker da ação no formato usado pelo Yahoo Finance.
    :return: Float, o preço atual da ação com duas casas decimais.
    """
    try:
        stock = yfinance.Ticker(ticker)
        preco_atual = stock.history(period="1d")['Close'].iloc[-1]
        return round(preco_atual, 2)
    except Exception as e:
        print(f"Erro: {e}")
        return None


# Create a new Excel file
wb = Workbook()
ws = wb.active
ws.title = Ticker

# Create 3 new sheets
ws1 = wb.create_sheet("Main")
ws2 = wb.create_sheet("Model")
ws3 = wb.create_sheet("Extra")

# Making the Main Sheet
ws1['A1'] = f'Equity Research {Ticker}, {datetime.today().strftime("%d/%m/%Y")}'
ws1['A1'].font = Font(bold=True, size=16)


ws1['B2'] = 'Company Name'
ws1['B3'] = 'Ticker'
ws1['B4'] = 'Sector'
ws1['B5'] = 'Industry'
ws1['B6'] = 'Founded'
ws1['B7'] = 'Headquarters'
ws1['B10'] = 'Beta'
ws1['B11'] = '52 Week High'
ws1['B12'] = '52 Week Low'
ws1['B13'] = '50 Day Moving Average'
ws1['B14'] = 'VaR 95%'
ws1['B15'] = 'Sharpe Ratio'



ws1['D2'] = stock.info['longName']
ws1['D3'] = stock.info['symbol']
ws1['D4'] = stock.info['sector']
ws1['D5'] = stock.info['industry']
ws1['D7'] = stock.info['city'] + ', ' + stock.info['country']
ws1['D10'] = calcular_beta(Ticker, benchmark)
ws1['D11'] = stock.info['fiftyTwoWeekHigh']
ws1['D12'] = stock.info['fiftyTwoWeekLow']
ws1['D13'] = stock.info['fiftyDayAverage']
ws1['D14'] = calcular_VaR(Ticker)
ws1['D15'] = calcular_sharpe_ratio(Ticker, rf)


ws1['B17'] = 'Description'
ws1['B17'].font = Font(bold=True)
ws1['B17'].alignment = Alignment(horizontal='left')
ws1['B17'].border = Border(bottom=Side(border_style='thin'))
ws1.merge_cells('B17:D17')
ws1['B18'] = stock.info['longBusinessSummary']
ws1['B18'].alignment = Alignment(horizontal='left')
ws1['B18'].border = Border(bottom=Side(border_style='thin'))


ws1['O2'] = 'Key Metrics'
ws1['O2'].font = Font(bold=True)
ws1['O2'].alignment = Alignment(horizontal='left')
ws1['O2'].border = Border(bottom=Side(border_style='thin'))
ws1.merge_cells('O2:P2')
ws1['O3'] = 'Price'
ws1['O4'] = 'Shares Outstanding'
ws1['O5'] = 'Market Cap'
ws1['O6'] = 'Cash'
ws1['O7'] = 'Debt'
ws1['O8'] = 'Enterprise Value'
ws1['O9'] = 'NPV'
ws1['O10'] = 'Fair Price'
ws1['O11'] = 'Change %'


# Today's closing Price
ws1['Q3'] = obter_preco_atual(Ticker)
ws1['Q4'] = stock.info['sharesOutstanding']
ws1['Q5'] = stock.info['marketCap']
ws1['Q6'] = stock.info['totalCash']
ws1['Q7'] = stock.info['totalDebt']
ws1['Q8'] = stock.info['enterpriseValue']
ws1['Q9'] = '?'
ws1['Q10'] = stock.info['fiftyTwoWeekHigh']
ws1['Q11'] = '=Q3/Q10-1'


# Set Column A to 4 width
ws1.column_dimensions['A'].width = 4

# Set column B to Bold and justified
for i in range(2, 8):
    ws1[f'B{i}'].font = Font(bold=True)
    ws1[f'B{i}'].alignment = Alignment(horizontal='left')
    
# Set column C to justified
for i in range(2, 8):
    ws1[f'D{i}'].alignment = Alignment(horizontal='left')


# Making Extras Sheet


# Analistas Recomendações
ws3['B2'] = 'Analysts Recommendations'
ws3.merge_cells('B2:D2')
ws3['B2'].font = Font(bold=True)
ws3['B2'].alignment = Alignment(horizontal='left')
ws3['B2'].border = Border(bottom=Side(border_style='thin'))
ws3['B3'] = 'Grade'
ws3['B3'].font = Font(bold=True)
ws3['B3'].alignment = Alignment(horizontal='left')
ws3['B3'].border = Border(bottom=Side(border_style='thin'))
ws3['C3'] = 'Firm'
ws3['C3'].font = Font(bold=True)
ws3['C3'].alignment = Alignment(horizontal='left')
ws3['C3'].border = Border(bottom=Side(border_style='thin'))
ws3['D3'] = 'Date'
ws3['D3'].font = Font(bold=True)
ws3['D3'].alignment = Alignment(horizontal='left')
ws3['D3'].border = Border(bottom=Side(border_style='thin'))


# Principais Executivos e suas Remunerações
ws3['B22'] = 'Key Executives'
ws3['B22'].font = Font(bold=True)
ws3['B22'].alignment = Alignment(horizontal='left')
ws3['B22'].border = Border(bottom=Side(border_style='thin'))
ws3.merge_cells('B22:F22')
ws3['B23'] = 'Name'
ws3['B23'].font = Font(bold=True)
ws3['B23'].alignment = Alignment(horizontal='left')
ws3['B23'].border = Border(bottom=Side(border_style='thin'))
ws3['C23'] = 'Title'
ws3['C23'].font = Font(bold=True)
ws3['C23'].alignment = Alignment(horizontal='left')
ws3['C23'].border = Border(bottom=Side(border_style='thin'))
ws3['D23'] = 'Pay'
ws3['D23'].font = Font(bold=True)
ws3['D23'].alignment = Alignment(horizontal='left')
ws3['D23'].border = Border(bottom=Side(border_style='thin'))

# Verificar se a chave 'companyOfficers' existe e tem conteúdo
if 'companyOfficers' in stock.info and stock.info['companyOfficers']:
    company_officers = stock.info['companyOfficers']
    row = 24
    for officer in company_officers:
        ws3[f'B{row}'] = officer.get('name', 'N/A')
        ws3[f'C{row}'] = officer.get('title', 'N/A')
        ws3[f'D{row}'] = officer.get('totalPay', 'N/A')
        row += 1
        if row > 28:  # Limitar aos 5 principais executivos
            break
    
    
ws3['B30'] = 'Sustainability and Corporate Governance'
ws3['B30'].font = Font(bold=True)
ws3['B30'].alignment = Alignment(horizontal='left')
ws3['B30'].border = Border(bottom=Side(border_style='thin'))
ws3.merge_cells('B30:D30')

# Dados de sustentabilidade (ESG)
ws3['B30'] = 'Sustainability and Corporate Governance'
ws3['B30'].font = Font(bold=True)
ws3['B30'].alignment = Alignment(horizontal='left')
ws3['B30'].border = Border(bottom=Side(border_style='thin'))
ws3.merge_cells('B30:D30')

# Tente obter dados de sustentabilidade
try:
    sustainability_data = stock.sustainability
    if sustainability_data is not None:
        row = 31
        for index, value in sustainability_data.iterrows():
            ws3[f'B{row}'] = index
            ws3[f'C{row}'] = value[0]
            row += 1
except yfinance.exceptions.YFNotImplementedError:
    print("Dados de sustentabilidade não disponíveis ou não implementados.")

        
ws3['B35'] = 'Major Shareholders'
ws3['B35'].font = Font(bold=True)
ws3['B35'].alignment = Alignment(horizontal='left')
ws3['B35'].border = Border(bottom=Side(border_style='thin'))
ws3.merge_cells('B35:D35')

# Dados dos principais acionistas
major_holders = stock.major_holders
major_holders = stock.major_holders
if major_holders is not None:
    # Imprimir as primeiras linhas do DataFrame para verificar sua estrutura
    print(major_holders.head())

    # Processar os dados dos principais acionistas
    row = 36
    # Verificar se as colunas esperadas estão presentes
    if 'Holder' in major_holders.columns and '% Out' in major_holders.columns:
        for index, value in major_holders.iterrows():
            ws3[f'B{row}'] = value['Holder']
            ws3[f'C{row}'] = value['% Out']
            row += 1


# Loop through recommendations
# Recomendações dos Analistas
# Resumo das Recomendações dos Analistas
ws3['B2'] = 'Analysts Recommendations Summary'
ws3.merge_cells('B2:F2')
ws3['B2'].font = Font(bold=True)
ws3['B2'].alignment = Alignment(horizontal='left')
ws3['B2'].border = Border(bottom=Side(border_style='thin'))
ws3['B3'] = 'Period'
ws3['B3'].font = Font(bold=True)
ws3['B3'].alignment = Alignment(horizontal='left')
ws3['B3'].border = Border(bottom=Side(border_style='thin'))
ws3['C3'] = 'Strong Buy'
ws3['C3'].font = Font(bold=True)
ws3['C3'].alignment = Alignment(horizontal='left')
ws3['C3'].border = Border(bottom=Side(border_style='thin'))
ws3['D3'] = 'Buy'
ws3['D3'].font = Font(bold=True)
ws3['D3'].alignment = Alignment(horizontal='left')
ws3['D3'].border = Border(bottom=Side(border_style='thin'))
ws3['E3'] = 'Hold'
ws3['E3'].font = Font(bold=True)
ws3['E3'].alignment = Alignment(horizontal='left')
ws3['E3'].border = Border(bottom=Side(border_style='thin'))
ws3['F3'] = 'Sell'
ws3['F3'].font = Font(bold=True)
ws3['F3'].alignment = Alignment(horizontal='left')
ws3['F3'].border = Border(bottom=Side(border_style='thin'))
ws3['G3'] = 'Strong Sell'
ws3['G3'].font = Font(bold=True)
ws3['G3'].alignment = Alignment(horizontal='left')
ws3['G3'].border = Border(bottom=Side(border_style='thin'))

# Obter o resumo das recomendações
recommendations_summary = stock.recommendations_summary
if recommendations_summary is not None:
    row = 4
    for index, summary in recommendations_summary.iterrows():
        ws3[f'B{row}'] = summary.get('period', 'N/A')
        ws3[f'C{row}'] = summary.get('strongBuy', 'N/A')
        ws3[f'D{row}'] = summary.get('buy', 'N/A')
        ws3[f'E{row}'] = summary.get('hold', 'N/A')
        ws3[f'F{row}'] = summary.get('sell', 'N/A')
        ws3[f'G{row}'] = summary.get('strongSell', 'N/A')
        row = row + 1

# Fulltime Employees
ws3['B18'] = 'Full Time Employees'
ws3['B18'].font = Font(bold=True)
ws3['C18'] = stock.info.get('fullTimeEmployees', 'N/A')


# Inserting images
ticker = yfinance.Ticker(Ticker)
hist = ticker.history(period="6y")

# Calcular a média móvel de 100 dias
hist['100d_ma'] = hist['Close'].rolling(window=100).mean()

# Criar o gráfico de preço e média móvel
plt.figure(figsize=(10, 6))
plt.plot(hist['Close'], label='Preço de Fechamento')
plt.plot(hist['100d_ma'], label='Média Móvel 100 dias', linestyle='--')
plt.title('Preço de Fechamento e Média Móvel 100 Dias - Últimos 6 Anos')
plt.legend()
plt.savefig('preco_e_media.png')
plt.close()

# Criar o gráfico de volume
plt.figure(figsize=(10, 4))
plt.bar(hist.index, hist['Volume'])
plt.title('Volume - Últimos 6 Anos')
plt.savefig('volume.png')
plt.close()



# Adicionando a imagem do gráfico na planilha
img1 = Image('preco_e_media.png')
img2 = Image('volume.png')

# Ajustar as propriedades da imagem conforme necessário
ws3.add_image(img1, 'O1') #
ws3.add_image(img2, 'O20') 

# Função para obter os preços de fechamento e dividend yields anuais
def obter_precos_dividend_yields(ticker, anos=5):
    agora = datetime.now()
    inicio = agora - timedelta(days=anos * 365)
    dados = yfinance.download(ticker, start=inicio.strftime('%Y-%m-%d'), end=agora.strftime('%Y-%m-%d'))

    # Verificar se a coluna 'Dividends' existe nos dados
    if 'Dividends' in dados.columns:
        # Obter os preços de fechamento
        precos_fechamento = dados['Close'].resample('Y').last()
        dividendos = dados['Dividends'].resample('Y').sum()

        # Calcular o dividend yield como total de dividendos pagos / preço de fechamento
        dividend_yield = dividendos / precos_fechamento.shift(1) * 100
    else:
        # Se a coluna 'Dividends' não existe, configurar os dividend yields como zero ou NaN
        precos_fechamento = dados['Close'].resample('Y').last()
        dividend_yield = pd.Series(np.nan, index=precos_fechamento.index)

    return precos_fechamento, dividend_yield

# Obter os preços de fechamento e dividend yields
precos_fechamento, dividend_yield = obter_precos_dividend_yields(Ticker)

# Configuração inicial da linha e coluna
row = 18  # Começando na linha 18
coluna_inicio = 'I'  # Começando na coluna I

# Adicionar cabeçalhos de coluna para os anos
for i, ano in enumerate(precos_fechamento.index.year):
    ws3[f'{chr(ord(coluna_inicio) + i)}{row}'] = ano

# Adicionar valores de fechamento e dividend yields
for i, (preco, dy) in enumerate(zip(precos_fechamento, dividend_yield)):
    # Preço de fechamento
    ws3[f'{chr(ord(coluna_inicio) + i)}{row + 1}'] = preco
    # Dividend yield
    ws3[f'{chr(ord(coluna_inicio) + i)}{row + 2}'] = f'{dy:.2f}%'
    




# making the model
ws2['A1'] = 'Main'


# Inserir borda grossa na parte inferior da linha 2
for col in range(1, 25):
    ws2.cell(row=2, column=col).border = Border(bottom=Side(border_style='thick'))


# Obter a demonstração de resultados trimestrais
quarterly_income_stmt = stock.quarterly_income_stmt

# Inverter a ordem das linhas e resetar o índice
quarterly_income_stmt = quarterly_income_stmt.iloc[::-1].reset_index()

# Renomear a coluna de índice para 'Account'
quarterly_income_stmt.rename(columns={'index': 'Account'}, inplace=True)

# Definir a linha inicial para inserção dos dados
current_row = 2  # Começa na linha 2 para inserir os nomes das contas
start_col = 2    # Começa na coluna B para inserir os valores

# Inserir os títulos das colunas (datas)
# As datas devem estar em ordem crescente, então primeiro ordenamos e depois inserimos
q_dates = [date for date in quarterly_income_stmt.columns if isinstance(date, datetime)]
q_dates.sort()

# Preencher as datas na linha 2 a partir da coluna B
for col, date in enumerate(q_dates, start=start_col):  # Começando da coluna B
    ws2.cell(row=current_row, column=col, value=date.strftime('%Y-%m-%d'))

# Inserir os dados de cada conta
for row_index, row in quarterly_income_stmt.iterrows():
    # Inserir o nome da conta na coluna A
    ws2.cell(row=current_row + row_index + 1, column=start_col - 1, value=row['Account'])
    for col_index, value in enumerate(row[1:]):
        ws2.cell(row=current_row + row_index + 1, column=start_col + col_index, value=value)


# Obter a demonstração de resultados
income_stmt = stock.income_stmt

# Inverter a ordem das linhas e resetar o índice
income_stmt = income_stmt.iloc[::-1].reset_index()

# Renomear a coluna de índice para 'Account'
income_stmt.rename(columns={'index': 'Account'}, inplace=True)

# Inserir os dados da demonstração de resultados na planilha
# Definir a linha inicial para inserção dos dados
current_row = 2  # Começa na linha 2
start_col = 12   # Começa na coluna L

# Inserir os títulos das colunas (datas)
# As datas devem estar em ordem crescente, então primeiro ordenamos e depois inserimos
dates = [date for date in income_stmt.columns if isinstance(date, datetime)]
dates.sort()

# Preencher as datas na linha 2 a partir da coluna L
for col, date in enumerate(dates, start=start_col):  # Começando da coluna L
    ws2.cell(row=current_row, column=col, value=date.strftime('%Y-%m-%d'))

# Inserir os dados de cada conta
for row_index, row in income_stmt.iterrows():
    # Inserir os dados começando da coluna L
    ws2.cell(row=current_row + row_index + 1, column=start_col - 1, value=row['Account'])
    for col_index, value in enumerate(row[1:]):
        ws2.cell(row=current_row + row_index + 1, column=start_col + col_index, value=value)



# Obter a demonstração de resultados anuais
# Obter a linha onde termina o income statement anual
last_row_income_stmt = ws2.max_row

# Adiciona 3 linhas de espaço conforme solicitado
start_row_quarterly_balance_sheet = last_row_income_stmt + 4

# Obter o balance sheet trimestral
quarterly_balance_sheet = stock.quarterly_balance_sheet

# Inverter a ordem das linhas e resetar o índice
quarterly_balance_sheet = quarterly_balance_sheet.iloc[::-1].reset_index()

# Renomear a coluna de índice para 'Account'
quarterly_balance_sheet.rename(columns={'index': 'Account'}, inplace=True)

# Inserir os títulos das colunas (datas) para o balance sheet trimestral
qb_dates = [date for date in quarterly_balance_sheet.columns if isinstance(date, datetime)]
qb_dates.sort()

# Preencher as datas na linha correspondente ao balance sheet trimestral a partir da coluna B
for col, date in enumerate(qb_dates, start=2):  # Começando da coluna B
    ws2.cell(row=start_row_quarterly_balance_sheet, column=col, value=date.strftime('%Y-%m-%d'))

# Inserir os dados de cada conta do balance sheet trimestral
for row_index, row in quarterly_balance_sheet.iterrows():
    ws2.cell(row=start_row_quarterly_balance_sheet + row_index + 1, column=1, value=row['Account'])
    for col_index, value in enumerate(row[1:]):
        ws2.cell(row=start_row_quarterly_balance_sheet + row_index + 1, column=col_index + 2, value=value)

# Calcular o número da próxima linha após os dados do balance sheet trimestral e o espaço adicional
next_row_after_quarterly_balance_sheet = start_row_quarterly_balance_sheet + len(quarterly_balance_sheet) + 3

# Obter o balance sheet anual
balance_sheet = stock.balance_sheet

# Inverter a ordem das linhas e resetar o índice
balance_sheet = balance_sheet.iloc[::-1].reset_index()

# Renomear a coluna de índice para 'Account'
balance_sheet.rename(columns={'index': 'Account'}, inplace=True)

# Inserir os títulos das colunas (datas) para o balance sheet anual
b_dates = [date for date in balance_sheet.columns if isinstance(date, datetime)]
b_dates.sort()

# Preencher as datas na linha correspondente ao balance sheet anual a partir da coluna L
for col, date in enumerate(b_dates, start=start_col):  # Começando da coluna L
    ws2.cell(row=next_row_after_quarterly_balance_sheet, column=col, value=date.strftime('%Y-%m-%d'))

# Inserir os dados de cada conta do balance sheet anual
for row_index, row in balance_sheet.iterrows():
    ws2.cell(row=next_row_after_quarterly_balance_sheet + row_index + 1, column=start_col - 1, value=row['Account'])
    for col_index, value in enumerate(row[1:]):
        ws2.cell(row=next_row_after_quarterly_balance_sheet + row_index + 1, column=start_col + col_index, value=value)


# Folder to save the Excel file
folder = r'C:\Users\conta\OneDrive\Área de Trabalho\Valuation'

# save the file
wb.save(folder + f'\\{Ticker}.xlsx')